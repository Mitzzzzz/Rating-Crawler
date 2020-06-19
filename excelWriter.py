import openpyxl
from openpyxl.styles import Alignment, Font
import json
import os

merged_items = []
imdbCrawlData = []
rottenCrawlData = []
keys_array = ["movieName",
              "yearofRelease",
              "Director",
              "languages",
              "genres",
              "imdbPlotSummary",
              "Cast",
              "imdbRating",
              "totalIMDBReviews",
              "lessthanorEqualtoFive",
              "greaterthanFive",
              "tomatometerScore",
              "tomatoAudienceScore",
              "totalRottenReviews",
              "lessthanorEqualtoThree",
              "greaterthanThree",
              "imdbURL",
              "rottenURL"]
try:
    with open("imdbCrawl.json", "r") as file:
        imdbCrawlData = json.load(file)
        print(imdbCrawlData)
except:
    pass
try:
    with open("rottenCrawl.json", "r") as file:
        rottenCrawlData = json.load(file)
        print(rottenCrawlData)
except:
    pass


def mergeItem(item_array):
    merged_string = ""
    for index, item in enumerate(item_array):
        if index == 0:
            merged_string = merged_string + item
        else:
            merged_string = merged_string + ", " + item
    return merged_string


def createNIL(item):
    for key in keys_array:
        if key not in item.keys() or item[key] == "null":
            if key in ["Director", "languages", "genres", "Cast"]:
                item[key] = ["NIL"]
            else:
                item[key] = "NIL"


def createObject(movie_object):
    obj = {"Movie Name": movie_object[0],
           "Year of Release": movie_object[1],
           "Director": movie_object[2],
           "Languages": movie_object[3],
           "Genre": movie_object[4],
           "Plot Summary": movie_object[5],
           "Cast": movie_object[6],
           "IMDB Rating": movie_object[7],
           "Total IMDB Reviews": movie_object[8],
           "IMDB Reviews Rated <=5": movie_object[9],
           "IMDB Reviews Rated >5": movie_object[10],
           "Tomatometer Score": movie_object[11],
           "Tomato Audience Score": movie_object[12],
           "Total RottenTomatoes Reviews": movie_object[13],
           "RottenTomatoes Reviews Rated <=3": movie_object[14],
           "RottenTomatoes Reviews Rated >3": movie_object[15],
           "IMDB URL": movie_object[16],
           "RottenTomatoes URL": movie_object[17]
           }
    merged_items.append(obj)


def standaloneRottenItem(rottenItem):
    createNIL(rottenItem)
    movie_obj = [rottenItem["movieName"],
                 rottenItem["yearofRelease"],
                 rottenItem["Director"],
                 mergeItem(rottenItem["languages"]),
                 mergeItem(rottenItem["genres"]),
                 rottenItem["imdbPlotSummary"],
                 mergeItem(rottenItem["Cast"]),
                 rottenItem["imdbRating"],
                 rottenItem["totalIMDBReviews"],
                 rottenItem["lessthanorEqualtoFive"],
                 rottenItem["greaterthanFive"],
                 rottenItem["tomatometerScore"],
                 rottenItem["tomatoAudienceScore"],
                 rottenItem["totalRottenReviews"],
                 rottenItem["lessthanorEqualtoThree"],
                 rottenItem["greaterthanThree"],
                 rottenItem["imdbURL"],
                 rottenItem["rottenURL"]
                 ]
    createObject(movie_obj)


def standaloneImdbItem(imdbItem):
    createNIL(imdbItem)
    movie_obj = [imdbItem["movieTitle"],
                 imdbItem["yearofRelease"],
                 mergeItem(imdbItem["Director"]),
                 mergeItem(imdbItem["languages"]),
                 mergeItem(imdbItem["genres"]),
                 imdbItem["imdbPlotSummary"],
                 mergeItem(imdbItem["Cast"]),
                 imdbItem["imdbRating"],
                 imdbItem["totalIMDBReviews"],
                 imdbItem["lessthanorEqualtoFive"],
                 imdbItem["greaterthanFive"],
                 imdbItem["tomatometerScore"],
                 imdbItem["tomatoAudienceScore"],
                 imdbItem["totalRottenReviews"],
                 imdbItem["lessthanorEqualtoThree"],
                 imdbItem["greaterthanThree"],
                 imdbItem["imdbURL"],
                 imdbItem["rottenURL"]
                 ]
    createObject(movie_obj)


if len(imdbCrawlData) == 0:
    for rottenItem in rottenCrawlData:
        standaloneRottenItem(rottenItem)

elif len(rottenCrawlData) == 0:
    for imdbItem in imdbCrawlData:
        standaloneImdbItem(imdbItem)

else:
    for imdbItem in imdbCrawlData:
        createNIL(imdbItem)
        for rottenItem in rottenCrawlData:
            createNIL(rottenItem)
            if imdbItem["movieName"] == rottenItem["movieName"] and mergeItem(imdbItem["Director"]) == rottenItem[
                "Director"]:
                print("Matched...................")
                movie_obj = [imdbItem["movieTitle"],
                             imdbItem["yearofRelease"],
                             mergeItem(imdbItem["Director"]),
                             mergeItem(imdbItem["languages"]),
                             mergeItem(imdbItem["genres"]),
                             imdbItem["imdbPlotSummary"],
                             mergeItem(imdbItem["Cast"]),
                             imdbItem["imdbRating"],
                             imdbItem["totalIMDBReviews"],
                             imdbItem["lessthanorEqualtoFive"],
                             imdbItem["greaterthanFive"],
                             rottenItem["tomatometerScore"],
                             rottenItem["tomatoAudienceScore"],
                             rottenItem["totalRottenReviews"],
                             rottenItem["lessthanorEqualtoThree"],
                             rottenItem["greaterthanThree"],
                             imdbItem["imdbURL"],
                             rottenItem["rottenURL"]
                             ]
                if "TV Series" in imdbItem["movieTitle"]:
                    movie_obj[1] = rottenItem["yearofRelease"]
                createObject(movie_obj)
                imdbCrawlData.remove(imdbItem)
                rottenCrawlData.remove(rottenItem)

    if len(imdbCrawlData) != 0:
        for imdbItem in imdbCrawlData:
            standaloneImdbItem(imdbItem)
    if len(rottenCrawlData) != 0:
        for rottenItem in rottenCrawlData:
            standaloneRottenItem(rottenItem)

print(merged_items)
wb = openpyxl.Workbook()  # Open a new workbook with a sheet names Sheet
sheet = wb["Sheet"]
for index, key in enumerate(merged_items[0].keys()):
    sheet.cell(row=1, column=index + 1).value = key
    sheet.cell(row=1, column=index + 1).font = Font(bold=True)
    sheet.cell(row=1, column=index + 1).alignment = Alignment(horizontal='center')

for i in range(0, len(merged_items)):
    for j, value in enumerate(merged_items[i]):
        sheet.cell(row=i + 2, column=j + 1).value = merged_items[i][value]
        sheet.cell(row=i + 2, column=j + 1).alignment = Alignment(wrap_text=True, horizontal='center')
        print("Value Written " + str(sheet.cell(row=i + 2, column=j + 1).value))
dims = {}
for row in sheet.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
for col, value in dims.items():
    if int(value) > 100:
        value = 100
    sheet.column_dimensions[col].width = value

os.remove("imdbCrawl.json")
os.remove("rottenCrawl.json")
wb.save("Movie.xlsx")
